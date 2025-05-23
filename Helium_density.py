import pandas as pd
from neqsim.thermo import TPflash, fluid
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import numpy as np
import os
from openpyxl import load_workbook



def to_scalar(x):
    # If x is a numpy.matrix, convert it to an ndarray first.
    if isinstance(x, np.matrix):
        x = np.asarray(x)
    # Ensure we have a numpy array and then flatten it.
    arr = np.asarray(x)
    return float(arr.flatten()[0])


def read_and_calculate_densities(file_path: str, sheet_name: str = 0) -> pd.DataFrame:
    """
    Reads an Excel file, extracts the necessary columns, calculates densities
    using NeqSim for three different models (PR, SRK, GERG2008, Vega),
    and writes the new density columns back into the same sheet in the first
    available empty columns.

    Expected columns in the Excel file:
      - "T[K]"
      - "p[MPa]"
      - "Experimental Density [kg/m^3]"
      - "HYSYS SRK"
      - "HYSYS PR"

    Parameters:
        file_path (str): Path to the Excel file.
        sheet_name (str or int): Sheet name or index (default is 0).

    Returns:
        pd.DataFrame: DataFrame with the original data plus NeqSim-calculated densities.
    """
    # Read the Excel file into a DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Define the expected columns
    expected_columns = [
        "T[K]",
        "p[MPa]",
        "Experimental Density [kg/m^3]",
        "HYSYS SRK",
        "HYSYS PR"
    ]
    missing_columns = [col for col in expected_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Missing columns in Excel file: {missing_columns}")

    # Extract and copy required columns
    data = df[expected_columns].copy()

    # Prepare lists to store NeqSim density calculations
    pr_densities = []
    srk_densities = []
    gerg2008_densities = []
    vega_densities = []

    # Loop over each row and calculate densities
    for _, row in data.iterrows():
        T = row["T[K]"]
        P = row["p[MPa]"]
        # PR model
        fluid_pr = fluid("pr")
        fluid_pr.addComponent("helium", 1, "mol/sec")
        fluid_pr.setTemperature(T, "K")
        fluid_pr.setPressure(P, "MPa")
        TPflash(fluid_pr)
        fluid_pr.initThermoProperties()
        pr_densities.append(to_scalar(fluid_pr.getDensity()))

        # SRK model (and derived GERG2008/Vega)
        fluid_srk = fluid("srk")
        fluid_srk.addComponent("helium", 1, "mol/sec")
        fluid_srk.setTemperature(T, "K")
        fluid_srk.setPressure(P, "MPa")
        TPflash(fluid_srk)
        fluid_srk.initThermoProperties()
        srk_densities.append(to_scalar(fluid_srk.getDensity()))
        # GERG2008 density from gas phase
        gerg2008_densities.append(to_scalar(
            fluid_srk.getPhase("gas").getDensity_GERG2008()
        ))
        # Vega density (phase index 0 assumed gas)
        vega_densities.append(to_scalar(
            fluid_srk.getPhase(0).getDensity_Vega()
        ))

    # Append calculated densities to DataFrame
    data["NeqSim PR Density"] = pr_densities
    data["NeqSim SRK Density"] = srk_densities
    data["NeqSim GERG2008 Density"] = gerg2008_densities
    data["NeqSim Vega Density"] = vega_densities

    # Open the workbook and select the sheet
    wb = load_workbook(file_path)
    if isinstance(sheet_name, int):
        sheet = wb[wb.sheetnames[sheet_name]]
    else:
        sheet = wb[sheet_name]

    # Determine where to start writing new columns (first empty header column)
    header_row = 1
    start_col = sheet.max_column + 1
    new_cols = [
        "NeqSim PR Density",
        "NeqSim SRK Density",
        "NeqSim GERG2008 Density",
        "NeqSim Vega Density"
    ]

    # Write headers for new columns
    for idx, col_name in enumerate(new_cols):
        sheet.cell(row=header_row, column=start_col + idx, value=col_name)

    # Write the density values under each new column
    for i, row in enumerate(data.itertuples(index=False, name=None), start=2):
        for j, value in enumerate(row[-4:]):  # last 4 entries are densities
            sheet.cell(row=i, column=start_col + j, value=float(value))

    # Save changes back to the same file
    wb.save(file_path)

    return data

def plot_relative_density_difference(data: pd.DataFrame, save_path):
    """
    Creates a scatter plot showing the relative density difference (in %)
    between model predictions and experimental density, with marker colors
    representing temperature (using the jet colormap). Markers are hollow (only outlines)
    except for the "x" marker which uses the default marker color to be visible, and the legend 
    is placed below the plot with black markers.
    """
    import matplotlib.pyplot as plt
    import numpy as np
    import matplotlib.lines as mlines

    # Work on a copy and compute relative differences.
    df = data.copy()
    exp = df["Experimental Density [kg/m^3]"]
    df["RD HYSYS SRK (%)"]       = 100 * (df["HYSYS SRK"] - exp) / exp
    df["RD HYSYS PR (%)"]        = 100 * (df["HYSYS PR"] - exp) / exp
    df["RD NeqSim SRK (%)"]      = 100 * (df["NeqSim SRK Density"] - exp) / exp
    df["RD NeqSim PR (%)"]       = 100 * (df["NeqSim PR Density"] - exp) / exp
    df["RD NeqSim GERG2008 (%)"] = 100 * (df["NeqSim GERG2008 Density"] - exp) / exp
    df["RD NeqSim Vega (%)"]     = 100 * (df["NeqSim Vega Density"] - exp) / exp

    # Round temperatures and exclude specific ones.
    df["T_rounded"] = np.round(df["T[K]"]).astype(int)
    df = df[~df["T_rounded"].isin([375, 475])]

    # Create the figure and axis.
    fig, ax = plt.subplots(figsize=(14, 8))
    plt.rcParams.update({'font.size': 16})
    
    # Define models with their respective relative density difference columns and marker styles.
    models = [
        ("Vega NeqSim",   "RD NeqSim Vega (%)",     'p'),
        ("GERG NeqSim",   "RD NeqSim GERG2008 (%)", 's'),
        ("SRK NeqSim",    "RD NeqSim SRK (%)",      'o'),
        ("PR NeqSim",     "RD NeqSim PR (%)",       '^'),
        ("SRK HYSYS",     "RD HYSYS SRK (%)",       'D'),
        ("PR HYSYS",      "RD HYSYS PR (%)",        'x')  # marker for 'x'
    ]
    marker_size = 100

    # Create normalization and colormap (jet) for temperature.
    norm = plt.Normalize(vmin=df["T_rounded"].min(), vmax=df["T_rounded"].max())
    cmap = plt.cm.jet

    # Plot each model’s data.
    for label, col, marker in models:
        colors = cmap(norm(df["T_rounded"]))
        if marker == 'x':
            ax.scatter(
                df["p[MPa]"],
                df[col],
                marker=marker,
                s=marker_size,
                c=colors,   # using the default color setting
                linewidths=2,
                alpha=0.8,
                label=label
            )
        else:
            ax.scatter(
                df["p[MPa]"],
                df[col],
                marker=marker,
                s=marker_size,
                facecolors='none',   # hollow markers
                edgecolors=colors,   # colored edges represent temperature
                linewidths=2,
                alpha=0.8,
                label=label
            )

    ax.set_xlabel("Pressure (MPa)", fontsize=16)
    ax.set_ylabel(r'$100(\rho_{\mathrm{model}} - \rho_{\mathrm{exp}})/\rho_{\mathrm{exp}}$', fontsize=20)
    ax.set_xlim(0, 40)
    ax.set_ylim(-2.5, 12)
    ax.grid(True)
    ax.tick_params(axis='both', which='major', labelsize=14)

    # Create custom proxy legend handles with black markers.
    # We use mlines.Line2D to create a marker with the desired properties.
    legend_order = ["Vega NeqSim", "GERG NeqSim", "SRK NeqSim", "PR NeqSim", "SRK HYSYS", "PR HYSYS"]
    legend_handles = []
    for label, col, marker in models:
        if marker == 'x':
            # For the 'x' marker, use a solid black marker.
            handle = mlines.Line2D(
                [], [], marker=marker, color='black', linestyle='None',
                markersize=10, markeredgewidth=2
            )
        else:
            # For other markers, display as hollow markers with black edges.
            handle = mlines.Line2D(
                [], [], marker=marker, color='black', linestyle='None',
                markersize=10, markerfacecolor='none', markeredgewidth=2
            )
        legend_handles.append(handle)

    # Filter the handles in the order specified by legend_order.
    # Here we assume models list order matches legend_order.
    ordered_handles = [legend_handles[i] for i, (lbl, _, _) in enumerate(models) if lbl in legend_order]

    ax.legend(
        ordered_handles, legend_order,
        loc='upper center',
        bbox_to_anchor=(0.5, -0.15),
        ncol=len(legend_order),
        fontsize=14,
        frameon=False
    )

    # Add a colorbar representing the temperature.
    sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, ax=ax)
    cbar.set_label("Temperature (K)", fontsize=16, labelpad=20)

    plt.tight_layout(rect=[0, 0.05, 1, 1])
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()






def compute_ARD_density_statistics(data: pd.DataFrame) -> dict:
    """
    Computes the ARD (Average Relative Deviation) statistics for each model.
    
    ARD is calculated as:
        100 * (Calculated Density - Experimental Density) / Experimental Density
        
    The function creates ARD columns for:
      - 'SRK_NeqSim'
      - 'PR_NeqSim'
      - 'GERG'
      - 'SRK_HYSYS'
      - 'PR_HYSYS'
      
    It then computes the mean ARD and the maximum absolute ARD for each model and prints the results.
    
    Parameters:
        data (pd.DataFrame): DataFrame containing at least the following columns:
            - 'Experimental Density [kg/m^3]'
            - 'NeqSim SRK Density'
            - 'NeqSim PR Density'
            - 'NeqSim GERG2008 Density'
            - 'HYSYS SRK'
            - 'HYSYS PR'
    
    Returns:
        dict: A dictionary with the computed statistics for each model.
    """
    # Calculate ARD values for each model.
    data['ARD SRK_NeqSim (%)'] = 100 * (data['NeqSim SRK Density'] - data['Experimental Density [kg/m^3]']) / data['Experimental Density [kg/m^3]']
    data['ARD PR_NeqSim (%)']  = 100 * (data['NeqSim PR Density'] - data['Experimental Density [kg/m^3]']) / data['Experimental Density [kg/m^3]']
    data['ARD GERG (%)']       = 100 * (data['NeqSim GERG2008 Density'] - data['Experimental Density [kg/m^3]']) / data['Experimental Density [kg/m^3]']
    data['ARD SRK_HYSYS (%)']  = 100 * (data['HYSYS SRK'] - data['Experimental Density [kg/m^3]']) / data['Experimental Density [kg/m^3]']
    data['ARD PR_HYSYS (%)']   = 100 * (data['HYSYS PR'] - data['Experimental Density [kg/m^3]']) / data['Experimental Density [kg/m^3]']
    data['ARD VEGA (%)']       = 100 * (data['NeqSim Vega Density'] - data['Experimental Density [kg/m^3]']) / data['Experimental Density [kg/m^3]']

    # Compute overall statistics.
    statistics = {
        'SRK_NeqSim': {
            'mean': data['ARD SRK_NeqSim (%)'].mean(),
            'max': data['ARD SRK_NeqSim (%)'].abs().max()
        },
        'PR_NeqSim': {
            'mean': data['ARD PR_NeqSim (%)'].mean(),
            'max': data['ARD PR_NeqSim (%)'].abs().max()
        },
        'GERG': {
            'mean': data['ARD GERG (%)'].mean(),
            'max': data['ARD GERG (%)'].abs().max()
        },
        'SRK_HYSYS': {
            'mean': data['ARD SRK_HYSYS (%)'].mean(),
            'max': data['ARD SRK_HYSYS (%)'].abs().max()
        },
        'PR_HYSYS': {
            'mean': data['ARD PR_HYSYS (%)'].mean(),
            'max': data['ARD PR_HYSYS (%)'].abs().max()
        },
        'VEGA' : {
            'mean': data['ARD VEGA (%)'].mean(),
            'max' : data['ARD VEGA (%)'].abs().max()
        }
    }

    # Print the results.
    print("Overall Results:")
    for model, stats in statistics.items():
        print(f"  {model} - Mean ARD (%): {stats['mean']:.5f}, Max ARD (%): {stats['max']:.5f}")

    return statistics




############### SPEED OF SOUND #####################################




def read_and_calculate_speed_of_sound(file_path: str, sheet_name: str = 0) -> pd.DataFrame:
    """
    Reads an Excel sheet, calculates speed of sound with four EOS in NeqSim,
    appends them under columns titled "<EOS> (NeqSim)", ensures they're floats,
    and writes back—overwriting only that sheet.
    """
    # 1) Figure out the real sheet name if they passed an index
    xls = pd.ExcelFile(file_path)
    if isinstance(sheet_name, int):
        sheet_name_str = xls.sheet_names[sheet_name]
    else:
        sheet_name_str = sheet_name

    # 2) Read the sheet
    df = pd.read_excel(file_path, sheet_name=sheet_name_str)
    req = ["T[K]", "p[MPa]", "Experimanetal c [m/s]"]
    missing = [c for c in req if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}")
    data = df[req].copy()

    # 3) Compute speeds
    eos_models = {
        "PR":  lambda f: to_scalar(f.getSoundSpeed()),
        "SRK": lambda f: to_scalar(f.getSoundSpeed()),
        "GERG2008": lambda f: to_scalar(f.getPhase(0).getProperties_GERG2008()[11]),
        "Vega":     lambda f: to_scalar(f.getPhase(0).getProperties_Vega()[11]),
    }

    # Prepare storage
    results = {name: [] for name in eos_models}

    for _, row in data.iterrows():
        T, P = row["T[K]"], row["p[MPa]"]
        # build two fluids, PR and SRK (GERG2008 & Vega piggy-back on the PR fluid)
        f_pr  = fluid("pr")
        f_srk = fluid("srk")
        for f in (f_pr, f_srk):
            f.addComponent("helium", 1, "mol/sec")
            f.setTemperature(T, "K")
            f.setPressure(P, "MPa")
            TPflash(f)
            f.initThermoProperties()

        # calculate and store
        results["PR"].append( float(eos_models["PR"](f_pr)) )
        results["SRK"].append(float(eos_models["SRK"](f_srk)))
        results["GERG2008"].append(float(eos_models["GERG2008"](f_pr)))
        results["Vega"].append(    float(eos_models["Vega"](f_pr)))

    # 4) Append columns, naming them "<EOS> (NeqSim)"
    for eos, vals in results.items():
        data[f"{eos} (NeqSim)"] = vals

    # 5) Ensure all the new columns are floats (catching any stray non-floats)
    neqsim_cols = [c for c in data.columns if c.endswith("(NeqSim)")]
    for c in neqsim_cols:
        data[c] = data[c].astype(float)

    # 6) Write back, overwriting only that one sheet
    with pd.ExcelWriter(file_path,
                        engine="openpyxl",
                        mode="a",
                        if_sheet_exists="replace") as writer:
        data.to_excel(writer, sheet_name=sheet_name_str, index=False)

    return data

def plot_relative_speed_of_sound_difference(data: pd.DataFrame, save_path):
    """
    Creates a single scatter plot showing the relative speed of sound difference (in %)
    between model predictions and experimental speed of sound across all isotherms.
    The relative difference is computed as:
        100 * (Calculated Speed of Sound - Experimental Speed of Sound) / Experimental Speed of Sound

    Markers are drawn with only outlines (no fill) using thicker lines. The legend is arranged in the following order:
        NeqSim Vega, NeqSim GERG2008, NeqSim SRK, NeqSim PR.
    The marker edge colors represent the isotherm temperature (using the jet colormap) and a colorbar is added.
    A single global y-axis label is added:
        100*(w_model - w_exp)/w_exp.
    """
    import matplotlib.pyplot as plt
    import numpy as np
    import matplotlib.lines as mlines

    # Use the default style and update font size.
    plt.style.use("default")
    plt.rcParams.update({'font.size': 16})
    
    # Work on a copy and compute relative differences.
    df = data.copy()
    exp = df["Experimanetal c [m/s]"]
    df["RD NeqSim SRK (%)"]      = 100 * (df["NeqSim SRK Speed of Sound"] - exp) / exp
    df["RD NeqSim PR (%)"]       = 100 * (df["NeqSim PR Speed of Sound"] - exp) / exp
    df["RD NeqSim GERG2008 (%)"] = 100 * (df["NeqSim GERG2008 Speed of Sound"] - exp) / exp
    df["RD NeqSim Vega (%)"]     = 100 * (df["NeqSim Vega Speed of Sound"] - exp) / exp

    # Group data by isotherm using rounded temperature values and exclude T = 398 K.
    df["T_rounded"] = np.round(df["T[K]"]).astype(int)
    df = df[df["T_rounded"] != 398]
    
    # Set up the figure and axis.
    fig, ax = plt.subplots(figsize=(14, 8))
    
    # Create normalization and colormap (jet) for temperature.
    norm = plt.Normalize(vmin=df["T_rounded"].min(), vmax=df["T_rounded"].max())
    cmap = plt.cm.jet

    # Define models in the desired legend order.
    models = [
        ("NeqSim Vega",     "RD NeqSim Vega (%)",     'p'),
        ("NeqSim GERG2008", "RD NeqSim GERG2008 (%)", 's'),
        ("NeqSim SRK",      "RD NeqSim SRK (%)",      'o'),
        ("NeqSim PR",       "RD NeqSim PR (%)",       '^')
    ]
    marker_size = 100

    # Plot each model's data, coloring marker edges by temperature.
    for label, col, marker in models:
        colors = cmap(norm(df["T_rounded"]))
        ax.scatter(
            df["p[MPa]"],
            df[col],
            marker=marker,
            s=marker_size,
            facecolors='none',  # hollow markers
            edgecolors=colors,  # marker edges colored by temperature
            linewidths=2,
            alpha=0.8,
            label=label
        )

    ax.set_xlabel("Pressure (MPa)", fontsize=16)
    ax.set_ylabel(r'$100\,(w_{\mathrm{model}} - w_{\mathrm{exp}})/w_{\mathrm{exp}}$', fontsize=20)
    ax.set_xlim(0, 16)
    ax.grid(True)
    ax.tick_params(axis='both', which='major', labelsize=14)

    # Create custom proxy legend handles with black markers.
    legend_order = ["NeqSim Vega", "NeqSim GERG2008", "NeqSim SRK", "NeqSim PR"]
    legend_handles = []
    for (label, col, marker) in models:
        # For hollow markers, we create a proxy using Line2D with markerfacecolor 'none'
        handle = mlines.Line2D(
            [], [], marker=marker, color='black', linestyle='None',
            markersize=10, markerfacecolor='none', markeredgewidth=2
        )
        legend_handles.append(handle)

    ax.legend(
        legend_handles, legend_order,
        loc='upper center',        # anchor legend at the upper center
        bbox_to_anchor=(0.5, -0.15),  # position it below the axes
        ncol=len(legend_order),
        fontsize=14,
        frameon=False,
        markerscale=1.2
    )

    # Add a colorbar representing the temperature.
    sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, ax=ax)
    cbar.set_label("Temperature (K)", fontsize=16, labelpad=20)
    
    # Adjust layout so there's enough space for the legend below.
    plt.tight_layout(rect=[0, 0.1, 1, 1])
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()


def isotherm_plot_relative_speed_of_sound_difference(data: pd.DataFrame, save_path):
    """
    Creates a grid of scatter plots, one per isotherm (excluding T = 398 K), showing the relative speed of sound
    difference (%) between model predictions and experimental values across pressure.
    Saves the resulting figure to the specified save_path.
    """
    import matplotlib.pyplot as plt
    import numpy as np
    import matplotlib.lines as mlines

    # Copy data and compute experimental reference
    df = data.copy()
    exp = df["Experimanetal c [m/s]"]

    # Compute relative differences for each model
    df["RD NeqSim SRK (%)"]      = 100 * (df["NeqSim SRK Speed of Sound"] - exp) / exp
    df["RD NeqSim PR (%)"]       = 100 * (df["NeqSim PR Speed of Sound"] - exp) / exp
    df["RD NeqSim GERG2008 (%)"] = 100 * (df["NeqSim GERG2008 Speed of Sound"] - exp) / exp
    df["RD NeqSim Vega (%)"]     = 100 * (df["NeqSim Vega Speed of Sound"] - exp) / exp

    # Round temperature and filter out unwanted isotherm
    df["T_rounded"] = np.round(df["T[K]"]).astype(int)
    df = df[df["T_rounded"] != 398]
    temps = sorted(df["T_rounded"].unique())
    plt.style.use("default")
    # Define models, markers and assign distinct colors
    color_cycle = ["blue", "blue", "blue", "blue"]
    models = [
        ("NeqSim Vega",     "RD NeqSim Vega (%)",     'p', color_cycle[0]),
        ("NeqSim GERG2008", "RD NeqSim GERG2008 (%)", 's', color_cycle[1]),
        ("NeqSim SRK",      "RD NeqSim SRK (%)",      'o', color_cycle[2]),
        ("NeqSim PR",       "RD NeqSim PR (%)",       '^', color_cycle[3])
    ]

    # Determine subplot grid size
    n_temps = len(temps)
    ncols = int(np.ceil(np.sqrt(n_temps)))
    nrows = int(np.ceil(n_temps / ncols))

    # Create subplots with larger figure size
    fig, axes = plt.subplots(nrows, ncols, figsize=(6 * ncols, 5 * nrows), sharex=True, sharey=True)
    axes = axes.flatten()

    # Plot each isotherm
    for ax, T in zip(axes, temps):
        sub = df[df["T_rounded"] == T]
        for label, col, marker, colr in models:
            ax.scatter(
                sub["p[MPa]"],
                sub[col],
                marker=marker,
                s=60,
                facecolors='none',  # hollow markers
                edgecolors=colr,
                linewidths=1.5,
                label=label
            )
        ax.set_title(f"{T} K", fontsize=14)
        ax.grid(True)
        ax.set_xlim(0, 1.55)

    # Remove any unused subplot axes
    for ax in axes[n_temps:]:
        fig.delaxes(ax)

    # Shared labels
    fig.supxlabel("Pressure (MPa)", fontsize=16)
    fig.supylabel(r'$100\,(w_{\mathrm{model}} - w_{\mathrm{exp}})/w_{\mathrm{exp}}$', fontsize=16)

    # Create legend handles with matching colors and place legend just below plots
    proxy_handles = []
    for label, _, marker, colr in models:
        proxy_handles.append(
            mlines.Line2D([], [], marker=marker, color=colr, linestyle='None',
                          markersize=8, markerfacecolor='none')
        )
    fig.legend(proxy_handles, [m[0] for m in models],
               loc='lower center', bbox_to_anchor=(0.5, -0.05),
               ncol=len(models), frameon=False, fontsize=16)

    # Tighten layout: less bottom margin
    plt.tight_layout(rect=[0, 0.05, 1, 1])
    fig.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close(fig)





def compute_ARD_speed_of_sound_statistics(data: pd.DataFrame) -> dict:
    exp = data['Experimanetal c [m/s]']
    # Use the new column names here:
    data['ARD SRK (NeqSim) (%)']      = 100 * (data['SRK (NeqSim)']      - exp) / exp
    data['ARD PR (NeqSim) (%)']       = 100 * (data['PR (NeqSim)']       - exp) / exp
    data['ARD GERG2008 (NeqSim) (%)'] = 100 * (data['GERG2008 (NeqSim)'] - exp) / exp
    data['ARD Vega (NeqSim) (%)']     = 100 * (data['Vega (NeqSim)']     - exp) / exp

    stats = {}
    for eos in ["SRK", "PR", "GERG2008", "Vega"]:
        col = f'ARD {eos} (NeqSim) (%)'
        stats[eos] = {
            'mean': data[col].mean(),
            'max':  data[col].abs().max()
        }

    print("Overall Results:")
    for eos, s in stats.items():
        print(f"  {eos}: Mean ARD = {s['mean']:.3f}%, Max ARD = {s['max']:.3f}%")
    return stats






if __name__ == "__main__":
    # Path to your Excel file containing the data.
    density_file_path = r"C:\Users\victo\OneDrive - NTNU\5. År\Master\Denisty_helium.xlsx"  # Update with your correct path
    speed_file_path = r"C:\Users\victo\OneDrive - NTNU\5. År\Master\speed_of_sound_helium_exp.xlsx"
    try:
        # Step 1 & 2: Read the Excel file and calculate NeqSim densities.
        density_data = read_and_calculate_densities(density_file_path, sheet_name=1)
        #print("Data with calculated densities:")
        #print(data.head())
        #print("Type of one of the calculated density values:",
        #      type(data["NeqSim GERG2008 Density"].iloc[0]))
        #print(data.dtypes)
        

        #stats = compute_ARD_density_statistics(density_data)

        save_path_density = os.path.join(r"C:\Users\victo\OneDrive - NTNU\5. År\Master\Helium_plots\Density", 
                             "Relative_density_difference_Ark2.png")
        #plot_relative_density_difference(density_data, save_path_density)
        speed_of_sound_data = read_and_calculate_speed_of_sound(speed_file_path, 3)

        save_path_speed = os.path.join(r"C:\Users\victo\OneDrive - NTNU\5. År\Master\Helium_plots\Speed of sound",
                                       "Relative_speed_of_sound_difference_Ark1.png")

        #plot_relative_speed_of_sound_difference(speed_of_sound_data, save_path_speed)
        stats = compute_ARD_speed_of_sound_statistics(speed_of_sound_data)

        #isotherm_plot_relative_speed_of_sound_difference(speed_of_sound_data,
         #                                                 save_path=os.path.join(r"C:\Users\victo\OneDrive - NTNU\5. År\Master\Helium_plots\Speed of sound",
          #                                                                       "Relative_speed_of_sound_difference_Ark2_isotherm.png"))
        
    except Exception as e:
        print(f"An error occurred: {e}")
